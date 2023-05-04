import tkinter as tk
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment

class Opositor:
    def __init__(self, nombre: str, apellidos: str, posicion: int, edad: int, ciudad_origen: str, ciudad_plaza: str, hospital_plaza: str, especialidad_deseada: str, ciudad_asignada: str, hospital_asignado: str, especialidad_asignada: str):
        self.nombre = nombre
        self.apellidos = apellidos
        self.posicion = posicion
        self.edad = edad
        self.ciudad_origen = ciudad_origen
        self.ciudad_plaza = ciudad_plaza
        self.hospital_plaza = hospital_plaza
        self.especialidad_deseada = especialidad_deseada
        self.ciudad_asignada = ciudad_asignada
        self.hospital_asignado = hospital_asignado
        self.especialidad_asignada = especialidad_asignada

    def update_plaza(self, ciudad_asignada, hospital_asignado, especialidad_asignada):
        self.ciudad_asignada = ciudad_asignada
        self.hospital_asignado = hospital_asignado
        self.especialidad_asignada = especialidad_asignada

    def update_info(self, nombre: str, apellidos: str, ciudad_plaza: str, hospital_plaza: str, especialidad_deseada: str):
        self.nombre = nombre
        self.apellidos = apellidos
        self.ciudad_plaza = ciudad_plaza
        self.hospital_plaza = hospital_plaza
        self.especialidad_deseada = especialidad_deseada

    def info1(self):
        return f"{self.nombre} {self.apellidos} - Posición: {self.posicion} - Plaza deseada: {self.ciudad_plaza}, {self.hospital_plaza}, {self.especialidad_deseada}"
    
    def info2(self):
        return f"- Plaza asignada: {self.ciudad_asignada}, {self.hospital_asignado}. {self.especialidad_asignada}"
    
    def delete_opositor(self):
        del self

class Plaza:
    def __init__(self, ciudad: str, hospital: str, especialidad: str, asignacion: int):
        self.ciudad = ciudad
        self.hospital = hospital
        self.especialidad = especialidad
        self.asignacion = asignacion
    
    def __str__(self):
        return f"{self.ciudad} {self.hospital} {self.especialidad} {self.asignacion}"
    
    def delete_plaza(self):
        del self

opositores = []
lista_plazas = []

try:
    lista_opositores = pd.read_excel("Lista_opositores.xlsx", header=1)
    for index, row in lista_opositores.iterrows():
        opositores.append(Opositor(row['Nombre'], row['Apellidos'], row['Posición'], row['Edad'], row['Ciudad_origen'], row['Ciudad'], row['Hospital'], row['Especialidad'], row['Ciudad:'], row['Hospital:'], row['Especialidad:']))
except FileNotFoundError:
    print("\nEl archivo Excel con la lista de opositores no está disponible. Asegúrate de copiarlo en el directorio desde el cual estás ejecutando este programa")
 
try:
    plazas = pd.read_excel("Plazas.xlsx",header=0)
    for index, row in plazas.iterrows():
        lista_plazas.append(Plaza(row['Ciudad'], row['Hospital'], row['Especialidad'], row['Asignacion']))
except FileNotFoundError:
    print("\nEl archivo Excel con la lista de plazas no está disponible. Asegúrate de copiarlo en el directorio desde el cual estás ejecutando este programa")
        
while True:

    a = int(input("\nIntroduce tu posición: "))

    for opositor in opositores:
        if opositor.posicion == a:
            opositor_actual = opositor
            posicion = opositores.index(opositor) 
            print(opositor_actual.info1())
            print(opositor_actual.info2())
    
    if opositor_actual.ciudad_plaza == opositor_actual.ciudad_asignada and opositor_actual.hospital_plaza == opositor_actual.hospital_asignado and opositor_actual.especialidad_deseada == opositor_actual.especialidad_asignada:
        continuar2 = input("\nTienes asignada la plaza que indicaste como más prioritaria. ¿Quieres cambiar de prioridad y comprobar si la nueva plaza está disponible? ")

        if continuar2 == "Si" or continuar2 == "si":
            print("\nIntroduce los datos de tu siguiente plaza más prioritaria: ")
            ciudad = input("\nCiudad: ")
            hospital = input("Hospital: ")
            especialidad = input("Especialidad: ")
            opositor_actual.update_info(opositor_actual.nombre, opositor_actual.apellidos, ciudad, hospital, especialidad)
            opositores[posicion]=opositor_actual

            for plaza in lista_plazas:
                if opositor_actual.ciudad_asignada == plaza.ciudad and opositor_actual.hospital_asignado == plaza.hospital and opositor_actual.especialidad_asignada == plaza.especialidad and opositor_actual.posicion == plaza.asignacion:
                    plaza.asignacion = 0
            continue
        else: break

    continuar = input("\n¿Es correcta esta información? ")
    if continuar == "Si" or continuar == "si" :
        for plaza in lista_plazas:
            if plaza.ciudad == opositor_actual.ciudad_plaza and plaza.hospital == opositor_actual.hospital_plaza and plaza.especialidad == opositor_actual.especialidad_deseada and (plaza.asignacion == 0 or plaza.asignacion > opositor_actual.posicion):

                respuesta = input("\nTu plaza está disponible. ¿Deseas que se te asigne? ")
                if respuesta == "si" or respuesta == "Si":
                    
                    for opositor in opositores:
                        if plaza.asignacion != 0 and opositor.posicion == plaza.asignacion:
                            opositor.ciudad_asignada = None
                            opositor.hospital_asignado = None
                            opositor.especialidad_asignada = None
                 
                    plaza.asignacion = opositor_actual.posicion
                    opositor_actual.update_plaza(plaza.ciudad, plaza.hospital, plaza.especialidad)
                    print("\n")
                    print(opositor_actual.info2())
                    print(plaza)

                    workbook = load_workbook(filename='Lista_opositores.xlsx')
                    worksheet = workbook.active
                    row = 3
                    for opositor in opositores:
                        worksheet.cell(row=row, column=9, value=opositor.ciudad_asignada).alignment = Alignment(horizontal='center')
                        worksheet.cell(row=row, column=10, value=opositor.hospital_asignado ).alignment = Alignment(horizontal='center')
                        worksheet.cell(row=row, column=11, value=opositor.especialidad_asignada).alignment = Alignment(horizontal='center')
                        row += 1
                    workbook.save(filename='Lista_opositores.xlsx')

                    workbook2 = load_workbook(filename='Plazas.xlsx')
                    worksheet2 = workbook2.active
                    row = 2
                    for plaza in lista_plazas:
                        worksheet2.cell(row=row, column=1, value=plaza.ciudad).alignment = Alignment(horizontal='center')
                        worksheet2.cell(row=row, column=2, value=plaza.hospital).alignment = Alignment(horizontal='center')
                        worksheet2.cell(row=row, column=3, value=plaza.especialidad).alignment = Alignment(horizontal='center')
                        worksheet2.cell(row=row, column=4, value=plaza.asignacion).alignment = Alignment(horizontal='center')
                        row += 1
                    workbook2.save(filename='Plazas.xlsx')

                    print("\nLos archivos con el listado de opositores y de plazas se han actualizado")
                else: continue

            elif plaza.ciudad == opositor_actual.ciudad_plaza and plaza.hospital == opositor_actual.hospital_plaza and plaza.especialidad == opositor_actual.especialidad_deseada and plaza.asignacion < opositor_actual.posicion:
                plaza_ocupada = True
                while plaza_ocupada:
                    print("\nTu plaza no está disponible ya que ha sido asignada al opositor con posición ",plaza.asignacion)
                    print("\nIntroduce los datos de tu siguiente plaza más prioritaria: ")
                    ciudad = input("\nCiudad: ")
                    hospital = input("Hospital: ")
                    especialidad = input("Especialidad: ")
                    opositor_actual.update_info(opositor_actual.nombre, opositor_actual.apellidos, ciudad, hospital, especialidad)
                    opositores[posicion]=opositor_actual # Creo que esto es reiterativo, en ningún momento hemos cambiado de opositor desde la línea 72
                    for plaza in lista_plazas:
                        if plaza.ciudad == opositor_actual.ciudad_plaza and plaza.hospital == opositor_actual.hospital_plaza and plaza.especialidad == opositor_actual.especialidad_deseada and (plaza.asignacion == 0 or plaza.asignacion > opositor_actual.posicion):
                            respuesta = input("\nTu plaza está disponible. ¿Deseas que se te asigne? ")
                            if respuesta == "si" or respuesta == "Si":

                                plaza.asignacion = opositor_actual.posicion
                                opositor_actual.update_plaza(plaza.ciudad, plaza.hospital, plaza.especialidad)
                                print("\n")
                                print(opositor_actual.info2())
                                print(plaza)

                                workbook = load_workbook(filename='Lista_opositores.xlsx')
                                worksheet = workbook.active
                                row = 3
                                for opositor in opositores:
                                    worksheet.cell(row=row, column=9, value=opositor.ciudad_asignada).alignment = Alignment(horizontal='center')
                                    worksheet.cell(row=row, column=10, value=opositor.hospital_asignado ).alignment = Alignment(horizontal='center')
                                    worksheet.cell(row=row, column=11, value=opositor.especialidad_asignada).alignment = Alignment(horizontal='center')
                                    row += 1
                                workbook.save(filename='Lista_opositores.xlsx')

                                workbook2 = load_workbook(filename='Plazas.xlsx')
                                worksheet2 = workbook2.active
                                row = 2
                                for plaza in lista_plazas:
                                    worksheet2.cell(row=row, column=1, value=plaza.ciudad).alignment = Alignment(horizontal='center')
                                    worksheet2.cell(row=row, column=2, value=plaza.hospital).alignment = Alignment(horizontal='center')
                                    worksheet2.cell(row=row, column=3, value=plaza.especialidad).alignment = Alignment(horizontal='center')
                                    worksheet2.cell(row=row, column=4, value=plaza.asignacion).alignment = Alignment(horizontal='center')
                                    row += 1
                                workbook2.save(filename='Plazas.xlsx')

                                plaza_ocupada = False
                                print("\nLos archivos con el listado de opositores y de plazas se han actualizado")

                            else: continue

            elif plaza.ciudad == opositor_actual.ciudad_plaza and plaza.hospital == opositor_actual.hospital_plaza and plaza.especialidad == opositor_actual.especialidad_deseada and plaza.asignacion == opositor_actual.posicion:
                print("\nEsta plaza ya se te ha asignado")
        break

    else: 
        ciudad = input("\nIntroduce los datos de la plaza que deseas:\nCiudad: ")
        hospital = input("Hospital: ")
        especialidad = input("Especialidad: ")
        opositor_actual.update_info(opositor_actual.nombre, opositor_actual.apellidos, ciudad, hospital, especialidad)
        opositores[posicion]=opositor_actual

        for plaza in lista_plazas:
                if opositor_actual.ciudad_asignada == plaza.ciudad and opositor_actual.hospital_asignado == plaza.hospital and opositor_actual.especialidad_asignada == plaza.especialidad and opositor_actual.posicion == plaza.asignacion:
                    plaza.asignacion = 0

        continue
